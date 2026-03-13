import os
import json
import uuid
import pandas as pd
from pathlib import Path
from django.shortcuts import render, redirect
from django.http import JsonResponse, FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings


UPLOAD_DIR = Path(settings.MEDIA_ROOT) / 'uploads'
OUTPUT_DIR = Path(settings.MEDIA_ROOT) / 'outputs'
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def index(request):
    request.session.flush()
    return render(request, 'merger/index.html')


def analyse(request):
    if request.method != 'POST':
        return redirect('index')

    file1 = request.FILES.get('file1')
    file2 = request.FILES.get('file2')

    if not file1 or not file2:
        return render(request, 'merger/index.html', {'error': 'Please upload both files.'})

    for f in [file1, file2]:
        if not f.name.endswith(('.xlsx', '.xls', '.csv')):
            return render(request, 'merger/index.html', {'error': f'Invalid file type: {f.name}. Only .xlsx, .xls, .csv allowed.'})

    session_id = str(uuid.uuid4())
    request.session['session_id'] = session_id

    def save_file(f):
        ext = Path(f.name).suffix
        fname = f'{session_id}_{f.name}'
        path = UPLOAD_DIR / fname
        with open(path, 'wb+') as dest:
            for chunk in f.chunks():
                dest.write(chunk)
        return str(path), f.name

    path1, name1 = save_file(file1)
    path2, name2 = save_file(file2)

    request.session['file1_path'] = path1
    request.session['file2_path'] = path2
    request.session['file1_name'] = name1
    request.session['file2_name'] = name2

    def read_file(path):
        if path.endswith('.csv'):
            return pd.read_csv(path)
        return pd.read_excel(path)

    try:
        df1 = read_file(path1)
        df2 = read_file(path2)
    except Exception as e:
        return render(request, 'merger/index.html', {'error': f'Error reading files: {str(e)}'})

    def analyse_df(df, name):
        info = {
            'name': name,
            'rows': len(df),
            'columns': list(df.columns),
            'col_count': len(df.columns),
            'dtypes': {col: str(df[col].dtype) for col in df.columns},
            'null_counts': {col: int(df[col].isnull().sum()) for col in df.columns},
            'sample': df.head(5).fillna('').to_dict(orient='records'),
            'duplicates': int(df.duplicated().sum()),
        }
        return info

    analysis1 = analyse_df(df1, name1)
    analysis2 = analyse_df(df2, name2)

    # Find common columns (potential merge keys)
    common_cols = list(set(df1.columns) & set(df2.columns))
    
    # Score columns by how useful they'd be as merge keys
    key_suggestions = []
    for col in common_cols:
        score = 0
        col_lower = col.lower()
        # Name-based scoring
        if any(k in col_lower for k in ['id', 'key', 'code', 'num', 'number', 'ref']):
            score += 3
        if any(k in col_lower for k in ['name', 'email', 'username', 'user']):
            score += 2
        # Uniqueness scoring
        u1 = df1[col].nunique() / max(len(df1), 1)
        u2 = df2[col].nunique() / max(len(df2), 1)
        score += (u1 + u2)
        key_suggestions.append({'col': col, 'score': round(score, 2)})
    
    key_suggestions.sort(key=lambda x: -x['score'])

    context = {
        'analysis1': analysis1,
        'analysis2': analysis2,
        'common_cols': common_cols,
        'key_suggestions': key_suggestions[:5],
        'all_cols1': list(df1.columns),
        'all_cols2': list(df2.columns),
        'analysis1_json': json.dumps(analysis1, default=str),
        'analysis2_json': json.dumps(analysis2, default=str),
    }
    return render(request, 'merger/analyse.html', context)


def merge(request):
    if request.method != 'POST':
        return redirect('index')

    path1 = request.session.get('file1_path')
    path2 = request.session.get('file2_path')
    name1 = request.session.get('file1_name', 'File 1')
    name2 = request.session.get('file2_name', 'File 2')

    if not path1 or not path2:
        return redirect('index')

    merge_key = request.POST.get('merge_key')
    merge_type = request.POST.get('merge_type', 'outer')
    handle_duplicates = request.POST.get('handle_duplicates', 'keep_first')
    output_name = request.POST.get('output_name', 'merged_output').strip() or 'merged_output'

    def read_file(path):
        if path.endswith('.csv'):
            return pd.read_csv(path)
        return pd.read_excel(path)

    try:
        df1 = read_file(path1)
        df2 = read_file(path2)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)

    if not merge_key or merge_key not in df1.columns or merge_key not in df2.columns:
        return JsonResponse({'error': f'Merge key "{merge_key}" not found in both files.'}, status=400)

    try:
        # Perform merge
        merged = pd.merge(df1, df2, on=merge_key, how=merge_type, suffixes=('_file1', '_file2'))

        # Handle duplicates
        if handle_duplicates == 'keep_first':
            merged = merged.drop_duplicates(subset=[merge_key], keep='first')
        elif handle_duplicates == 'keep_last':
            merged = merged.drop_duplicates(subset=[merge_key], keep='last')
        elif handle_duplicates == 'keep_all':
            pass  # keep everything

        # Save output
        safe_name = ''.join(c for c in output_name if c.isalnum() or c in ' _-').strip()
        session_id = request.session.get('session_id', 'output')
        out_filename = f'{session_id}_{safe_name}.xlsx'
        out_path = OUTPUT_DIR / out_filename

        with pd.ExcelWriter(str(out_path), engine='openpyxl') as writer:
            merged.to_excel(writer, sheet_name='Merged', index=False)
            df1.to_excel(writer, sheet_name=name1[:31], index=False)
            df2.to_excel(writer, sheet_name=name2[:31], index=False)

            # Style the merged sheet
            wb = writer.book
            ws = wb['Merged']
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_fill = PatternFill('solid', start_color='1E3A5F')
            header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
            alt_fill = PatternFill('solid', start_color='F0F4F8')
            border = Border(
                bottom=Side(style='thin', color='D0D7DE'),
            )
            for col_idx, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[1].height = 30

            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical='center')
                    if row_idx % 2 == 0:
                        cell.fill = alt_fill

            for col in ws.columns:
                max_len = max((len(str(c.value)) if c.value else 0) for c in col)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 40)

        stats = {
            'total_rows': len(merged),
            'columns': len(merged.columns),
            'file1_rows': len(df1),
            'file2_rows': len(df2),
            'filename': out_filename,
            'output_name': f'{safe_name}.xlsx',
        }
        request.session['output_file'] = out_filename
        return JsonResponse({'success': True, 'stats': stats, 'filename': out_filename})

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


def download(request, filename):
    # Security: only allow files from this session
    session_output = request.session.get('output_file', '')
    if filename != session_output:
        raise Http404

    file_path = OUTPUT_DIR / filename
    if not file_path.exists():
        raise Http404

    response = FileResponse(open(file_path, 'rb'),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    safe_name = filename.split('_', 2)[-1] if '_' in filename else filename
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response


def reset(request):
    request.session.flush()
    return redirect('index')
